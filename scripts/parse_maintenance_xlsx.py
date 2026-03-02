#!/usr/bin/env python3
"""Parse maintenance xlsx files and generate PDFs.

Reads 40 work-order xlsx files + 1 tool list, parses via label-based
extraction, exports unified JSON, and generates print-ready PDFs with
Korean font support.

Usage:
    python3 scripts/parse_maintenance_xlsx.py
"""

import glob
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from fpdf import FPDF

# ──────────────────────────────────────────────
# Paths
# ──────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data" / "정비이력_비용분석_ISO14224_연계_v7"
JSON_OUT = BASE_DIR / "data" / "parsed_maintenance_records.json"
PDF_OUT_DIR = BASE_DIR / "knowledge_base" / "maintenance_history"
EQUIP_PDF_DIR = BASE_DIR / "knowledge_base" / "equipment_manual"
FONT_PATH = str(BASE_DIR / "fonts" / "NotoSansKR.ttf")


def parse_instruction_sheet(ws: Any) -> dict[str, Any]:
    """Parse 지시서 (work order instruction) sheet by label matching."""
    data: dict[str, Any] = {}
    checklist: list[str] = []
    materials: list[dict[str, str]] = []
    tools: list[dict[str, str]] = []
    post_checks: list[str] = []
    attachments: list[str] = []

    in_materials = False
    in_tools = False

    for row in range(1, ws.max_row + 1):
        a_val = ws.cell(row, 1).value
        b_val = ws.cell(row, 2).value

        if a_val is None and b_val is None:
            in_materials = False
            in_tools = False
            continue

        a_str = str(a_val).strip() if a_val else ""
        b_str = str(b_val).strip() if b_val else ""

        # Header row
        if "작업지시서" in a_str and row == 1:
            continue

        # Material/tool table header detection
        if a_str == "자재 코드":
            in_materials = True
            in_tools = False
            continue
        if a_str == "공구/장비 코드":
            in_tools = True
            in_materials = False
            continue

        # Material/tool data rows
        if in_materials and a_str and a_str not in ("필요 자재",):
            materials.append({
                "code": a_str,
                "name": b_str,
                "spec": str(ws.cell(row, 3).value or ""),
                "qty": str(ws.cell(row, 4).value or ""),
                "unit": str(ws.cell(row, 5).value or ""),
                "note": str(ws.cell(row, 6).value or ""),
            })
            continue
        if in_tools and a_str and a_str not in ("필요 공구 및 장비",):
            tools.append({
                "code": a_str,
                "name": b_str,
                "spec": str(ws.cell(row, 3).value or ""),
                "qty": str(ws.cell(row, 4).value or ""),
                "unit": str(ws.cell(row, 5).value or ""),
                "note": str(ws.cell(row, 6).value or ""),
            })
            continue

        in_materials = False
        in_tools = False

        # Label-based field extraction
        if a_str == "작업지시 번호":
            data["wo_number"] = b_str
        elif a_str == "설비명 / 설비번호":
            data["equipment"] = b_str
        elif a_str == "설비 위치":
            data["location"] = b_str
        elif a_str == "작업 유형":
            data["work_type"] = b_str
        elif a_str == "작업 요청일":
            data["request_date"] = normalize_date(b_str)
        elif a_str == "작업 예정일":
            data["scheduled_date"] = normalize_datetime(b_str)
        elif a_str == "작업 완료 예정일":
            data["due_date"] = normalize_datetime(b_str)
        elif a_str == "작업 담당자":
            data["assignee"] = b_str
        elif a_str == "작업 내용 요약":
            data["summary"] = b_str
        elif a_str == "작업 안전사항":
            data["safety"] = b_str
        elif a_str == "작업 상세 내용(체크리스트)":
            checklist.append(b_str)
        elif a_str == "필요 자재":
            in_materials = True
        elif a_str == "필요 공구 및 장비":
            in_tools = True
        elif a_str == "작업 후 확인사항":
            post_checks.append(b_str)
        elif a_str == "작업 승인자":
            data["approver"] = b_str
        elif a_str == "작업 완료일시":
            data["completion_date"] = normalize_datetime(b_str)
        elif a_str == "작업 결과 요약":
            data["result_summary"] = b_str
        elif a_str == "첨부파일":
            attachments.append(b_str)

    data["checklist"] = checklist
    data["materials"] = materials
    data["tools"] = tools
    data["post_checks"] = post_checks
    data["attachments"] = attachments
    return data


def parse_completion_sheet(ws: Any) -> dict[str, Any]:
    """Parse 완료 (completion report) sheet by label matching."""
    data: dict[str, Any] = {}
    failure_causes: list[str] = []
    solutions: list[str] = []
    results: list[str] = []
    attachments: list[str] = []

    for row in range(1, ws.max_row + 1):
        a_val = ws.cell(row, 1).value
        b_val = ws.cell(row, 2).value
        if a_val is None:
            continue

        a_str = str(a_val).strip()
        b_str = str(b_val).strip() if b_val else ""

        if "완료 보고서" in a_str and row == 1:
            continue

        if a_str == "작업지시 번호":
            data["wo_number"] = b_str
        elif a_str == "설비명 / 설비번호":
            data["equipment"] = b_str
        elif a_str == "설비 위치":
            data["location"] = b_str
        elif a_str == "작업 유형":
            data["work_type"] = b_str
        elif a_str == "작업 시작":
            data["start_time"] = normalize_datetime(b_str)
        elif a_str == "작업 종료":
            data["end_time"] = normalize_datetime(b_str)
        elif a_str == "참여 인력":
            data["personnel"] = b_str
        elif a_str == "고장 원인(상세)":
            failure_causes.append(b_str)
        elif a_str == "해결 방법(작업 상세)":
            solutions.append(b_str)
        elif a_str == "조치 후 결과(측정값/판정)":
            results.append(b_str)
        elif a_str == "첨부파일":
            attachments.append(b_str)
        elif a_str == "보고 승인자":
            data["report_approver"] = b_str

    data["failure_causes"] = failure_causes
    data["solutions"] = solutions
    data["results"] = results
    data["attachments"] = attachments
    return data


def normalize_date(val: str) -> str:
    """Normalize date string to YYYY-MM-DD."""
    if not val or val == "None":
        return ""
    val = val.strip()
    # Already in YYYY-MM-DD
    if re.match(r"^\d{4}-\d{2}-\d{2}$", val):
        return val
    # datetime object string
    try:
        dt = datetime.fromisoformat(val)
        return dt.strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        return val


def normalize_datetime(val: str) -> str:
    """Normalize datetime string to YYYY-MM-DD HH:MM."""
    if not val or val == "None":
        return ""
    val = val.strip()
    if re.match(r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}$", val):
        return val
    try:
        dt = datetime.fromisoformat(val)
        return dt.strftime("%Y-%m-%d %H:%M")
    except (ValueError, TypeError):
        return val


def normalize_work_type(wt: str) -> str:
    """Normalize work type labels."""
    mapping = {
        "분석점검": "분석점검",
        "예방정비": "예방정비",
        "긴급수리": "긴급수리",
        "교체작업": "교체작업",
        "정기점검": "정기점검",
    }
    return mapping.get(wt.strip(), wt.strip())


def extract_line_product(location: str) -> tuple[str, str]:
    """Extract line and product from location string like '라인 A / 제품 AX-01'."""
    m = re.search(r"라인\s*(\w+)\s*/\s*제품\s*([\w-]+)", location)
    if m:
        return m.group(1), m.group(2)
    return "X", "XX-00"


def parse_all_files() -> list[dict[str, Any]]:
    """Parse all 40 DTL xlsx files."""
    files = sorted(glob.glob(str(DATA_DIR / "작업지시서_*.xlsx")))
    records: list[dict[str, Any]] = []
    quality_issues: list[str] = []

    for f in files:
        fname = os.path.basename(f)
        wb = openpyxl.load_workbook(f)

        instruction = {}
        completion = {}
        for sn in wb.sheetnames:
            ws = wb[sn]
            if "지시서" in sn:
                instruction = parse_instruction_sheet(ws)
            elif "완료" in sn:
                completion = parse_completion_sheet(ws)

        # Normalize
        if instruction.get("work_type"):
            instruction["work_type"] = normalize_work_type(instruction["work_type"])
        if completion.get("work_type"):
            completion["work_type"] = normalize_work_type(completion["work_type"])

        # Quality checks
        wo = instruction.get("wo_number", "")
        if not wo:
            quality_issues.append(f"{fname}: missing wo_number")
        if not instruction.get("equipment"):
            quality_issues.append(f"{fname}: missing equipment")
        if not completion.get("failure_causes"):
            quality_issues.append(f"{fname}: no failure causes in completion")

        records.append({
            "source_file": fname,
            "instruction": instruction,
            "completion": completion,
        })

    if quality_issues:
        print(f"[Quality] {len(quality_issues)} issues found:")
        for qi in quality_issues:
            print(f"  - {qi}")
    else:
        print("[Quality] All records passed validation.")

    return records


def parse_tool_list() -> list[dict[str, Any]]:
    """Parse tool/equipment list xlsx."""
    f = str(DATA_DIR / "공구 및 장비_품목리스트_V1.xlsx")
    wb = openpyxl.load_workbook(f)
    ws = wb["공구 및 장비품목"]
    items: list[dict[str, Any]] = []
    for row in range(2, ws.max_row + 1):
        items.append({
            "code": str(ws.cell(row, 1).value or ""),
            "name": str(ws.cell(row, 2).value or ""),
            "spec": str(ws.cell(row, 3).value or ""),
            "qty": ws.cell(row, 4).value,
            "unit": str(ws.cell(row, 5).value or ""),
            "lead_time_days": ws.cell(row, 6).value,
            "vendor": str(ws.cell(row, 7).value or ""),
            "unit_price_krw": ws.cell(row, 8).value,
        })
    return items


# ──────────────────────────────────────────────
# PDF Generation
# ──────────────────────────────────────────────
class MaintenancePDF(FPDF):
    """Custom PDF with Korean font support and proper text wrapping."""

    LABEL_W = 45  # Label column width
    LINE_H = 5.5  # Default line height
    PAGE_W = 210  # A4 width
    MARGIN = 10  # Left/right margin

    def __init__(self) -> None:
        super().__init__()
        self.add_font("NotoSansKR", "", FONT_PATH)
        self.add_font("NotoSansKR", "B", FONT_PATH)
        self.set_auto_page_break(auto=True, margin=15)
        self.set_left_margin(self.MARGIN)
        self.set_right_margin(self.MARGIN)

    @property
    def content_w(self) -> float:
        """Usable content width."""
        return self.PAGE_W - self.l_margin - self.r_margin

    def header(self) -> None:
        pass

    def section_title(self, title: str) -> None:
        self._check_space(20)  # Section title + at least one field
        self.set_font("NotoSansKR", "B", 11)
        self.set_fill_color(230, 230, 230)
        self.cell(self.content_w, 8, f"  {title}", ln=True, fill=True)
        self.ln(2)

    def _check_space(self, min_h: float = 12.0) -> None:
        """Add a new page if remaining space is less than min_h."""
        if self.get_y() + min_h > self.h - self.b_margin:
            self.add_page()

    def field(self, label: str, value: str) -> None:
        """Render a label-value pair with proper wrapping for long values."""
        value = value or "-"

        # Estimate height needed (prevent label/value split across pages)
        self.set_font("NotoSansKR", "", 9)
        val_w = self.content_w - self.LABEL_W
        # Rough estimate: fpdf2 get_string_width for multi-line
        lines = max(1, len(value) / max(val_w / 2.5, 1))
        est_h = max(self.LINE_H, lines * self.LINE_H)
        self._check_space(min(est_h + 2, 40))  # Cap at 40mm to avoid infinite page adds

        y_before = self.get_y()
        x0 = self.l_margin

        # Temporarily disable auto page break for this field
        self.set_auto_page_break(auto=False)

        # Label (bold, fixed width)
        self.set_font("NotoSansKR", "B", 9)
        self.set_xy(x0, y_before)
        self.cell(self.LABEL_W, self.LINE_H, label)

        # Value (regular, wrapping in remaining width)
        self.set_font("NotoSansKR", "", 9)
        self.set_xy(x0 + self.LABEL_W, y_before)
        self.multi_cell(val_w, self.LINE_H, value)

        # Re-enable auto page break
        self.set_auto_page_break(auto=True, margin=15)

        # Ensure cursor is below both label and value
        y_after = self.get_y()
        if y_after < y_before + self.LINE_H:
            self.set_y(y_before + self.LINE_H)

    def items_list(self, items: list[str]) -> None:
        """Render a bullet list with full-width wrapping."""
        self.set_font("NotoSansKR", "", 9)
        indent = 5
        for item in items:
            self._check_space(self.LINE_H * 2)
            self.set_x(self.l_margin + indent)
            w = self.content_w - indent
            self.multi_cell(w, self.LINE_H, f"• {item}")
            self.ln(0.5)

    def small_table(
        self,
        headers: list[str],
        rows: list[list[str]],
        col_widths: list[float],
    ) -> None:
        """Render a table with wrapped cell content."""
        row_h = 6

        # Header
        self.set_font("NotoSansKR", "B", 8)
        for i, h in enumerate(headers):
            self.cell(col_widths[i], row_h, h, border=1, align="C")
        self.ln()

        # Data rows — measure multi_cell height, then render
        self.set_font("NotoSansKR", "", 8)
        for row_data in rows:
            # Calculate max height needed for this row
            max_lines = 1
            for i, val in enumerate(row_data):
                # Estimate lines needed
                char_w = col_widths[i] / 4.5  # rough chars per line at font 8
                lines = max(1, len(val) / max(char_w, 1))
                max_lines = max(max_lines, lines)
            cell_h = max(row_h, int(max_lines) * row_h)

            # Check page break
            if self.get_y() + cell_h > self.h - self.b_margin:
                self.add_page()
                # Re-print header
                self.set_font("NotoSansKR", "B", 8)
                for i, h in enumerate(headers):
                    self.cell(col_widths[i], row_h, h, border=1, align="C")
                self.ln()
                self.set_font("NotoSansKR", "", 8)

            y0 = self.get_y()
            x0 = self.l_margin
            actual_max_y = y0

            for i, val in enumerate(row_data):
                self.set_xy(x0, y0)
                self.multi_cell(col_widths[i], row_h, val, border=0)
                actual_max_y = max(actual_max_y, self.get_y())
                x0 += col_widths[i]

            # Draw borders for each cell at the actual height
            actual_h = actual_max_y - y0
            if actual_h < row_h:
                actual_h = row_h
            x0 = self.l_margin
            for i in range(len(row_data)):
                self.rect(x0, y0, col_widths[i], actual_h)
                x0 += col_widths[i]

            self.set_y(y0 + actual_h)


def generate_wo_pdf(record: dict[str, Any], out_path: Path) -> None:
    """Generate a 2-page PDF for one work order."""
    inst = record["instruction"]
    comp = record["completion"]

    pdf = MaintenancePDF()

    # ── Page 1: 작업지시서 ──
    pdf.add_page()
    pdf.set_font("NotoSansKR", "B", 14)
    pdf.cell(0, 10, "공장설비 보전작업 작업지시서", ln=True, align="C")
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(4)

    pdf.section_title("기본 정보")
    pdf.field("작업지시 번호", inst.get("wo_number", ""))
    pdf.field("설비명 / 설비번호", inst.get("equipment", ""))
    pdf.field("설비 위치", inst.get("location", ""))
    pdf.field("작업 유형", inst.get("work_type", ""))
    pdf.field("작업 요청일", inst.get("request_date", ""))
    pdf.field("작업 예정일", inst.get("scheduled_date", ""))
    pdf.field("완료 예정일", inst.get("due_date", ""))
    pdf.field("작업 담당자", inst.get("assignee", ""))

    pdf.ln(2)
    pdf.section_title("작업 내용")
    pdf.field("작업 내용 요약", inst.get("summary", ""))
    pdf.field("안전사항", inst.get("safety", ""))

    if inst.get("checklist"):
        pdf.ln(2)
        pdf.section_title("작업 상세 체크리스트")
        pdf.items_list(inst["checklist"])

    if inst.get("materials"):
        pdf.ln(2)
        pdf.section_title("필요 자재")
        widths = [28.0, 32.0, 32.0, 14.0, 14.0, 70.0]
        headers = ["코드", "자재명", "규격", "수량", "단위", "비고"]
        rows = [[m["code"], m["name"], m["spec"], m["qty"], m["unit"], m["note"]] for m in inst["materials"]]
        pdf.small_table(headers, rows, widths)

    if inst.get("tools"):
        pdf.ln(2)
        pdf.section_title("필요 공구 및 장비")
        widths = [28.0, 32.0, 38.0, 14.0, 14.0, 64.0]
        headers = ["코드", "명칭", "규격", "수량", "단위", "비고"]
        rows = [[t["code"], t["name"], t["spec"], t["qty"], t["unit"], t["note"]] for t in inst["tools"]]
        pdf.small_table(headers, rows, widths)

    if inst.get("post_checks"):
        pdf.ln(2)
        pdf.section_title("작업 후 확인사항")
        pdf.items_list(inst["post_checks"])

    pdf.ln(2)
    pdf.section_title("승인 및 결과")
    pdf.field("작업 승인자", inst.get("approver", ""))
    pdf.field("작업 완료일시", inst.get("completion_date", ""))
    pdf.field("작업 결과 요약", inst.get("result_summary", ""))

    if inst.get("attachments"):
        pdf.field("첨부파일", ", ".join(inst["attachments"]))

    # ── Page 2: 완료보고서 ──
    pdf.add_page()
    pdf.set_font("NotoSansKR", "B", 14)
    pdf.cell(0, 10, "공장설비 보전작업 완료 보고서", ln=True, align="C")
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(4)

    pdf.section_title("기본 정보")
    pdf.field("작업지시 번호", comp.get("wo_number", ""))
    pdf.field("설비명 / 설비번호", comp.get("equipment", ""))
    pdf.field("설비 위치", comp.get("location", ""))
    pdf.field("작업 유형", comp.get("work_type", ""))
    pdf.field("작업 시작", comp.get("start_time", ""))
    pdf.field("작업 종료", comp.get("end_time", ""))
    pdf.field("참여 인력", comp.get("personnel", ""))

    if comp.get("failure_causes"):
        pdf.ln(2)
        pdf.section_title("고장 원인 (상세)")
        pdf.items_list(comp["failure_causes"])

    if comp.get("solutions"):
        pdf.ln(2)
        pdf.section_title("해결 방법 (작업 상세)")
        pdf.items_list(comp["solutions"])

    if comp.get("results"):
        pdf.ln(2)
        pdf.section_title("조치 후 결과 (측정값/판정)")
        pdf.items_list(comp["results"])

    if comp.get("attachments"):
        pdf.ln(2)
        pdf.field("첨부파일", ", ".join(comp["attachments"]))

    pdf.ln(2)
    pdf.field("보고 승인자", comp.get("report_approver", ""))

    pdf.output(str(out_path))


def generate_tool_list_pdf(tools: list[dict[str, Any]], out_path: Path) -> None:
    """Generate PDF for tool/equipment list."""
    pdf = MaintenancePDF()
    pdf.add_page()
    pdf.set_font("NotoSansKR", "B", 14)
    pdf.cell(0, 10, "공구 및 장비 품목리스트 V1", ln=True, align="C")
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(6)

    widths = [22.0, 28.0, 40.0, 12.0, 12.0, 16.0, 36.0, 24.0]
    headers = ["품목코드", "품명", "규격/모델", "수량", "단위", "리드타임", "벤더", "단가(KRW)"]
    rows: list[list[str]] = []
    for t in tools:
        price = f"{t['unit_price_krw']:,}" if t.get("unit_price_krw") else ""
        rows.append([
            t["code"], t["name"], t["spec"],
            str(t.get("qty", "")), t["unit"],
            str(t.get("lead_time_days", "")),
            t["vendor"], price,
        ])
    pdf.small_table(headers, rows, widths)

    pdf.output(str(out_path))


def main() -> None:
    """Main entry point."""
    print("=== Parsing maintenance xlsx files ===")
    records = parse_all_files()
    print(f"Parsed {len(records)} work order records.")

    # Parse tool list
    tool_items = parse_tool_list()
    print(f"Parsed {len(tool_items)} tool/equipment items.")

    # Save JSON
    JSON_OUT.parent.mkdir(parents=True, exist_ok=True)
    output_data = {
        "generated_at": datetime.now().isoformat(),
        "total_records": len(records),
        "records": records,
        "tool_list": tool_items,
    }
    with open(JSON_OUT, "w", encoding="utf-8") as f:
        json.dump(output_data, f, ensure_ascii=False, indent=2)
    print(f"JSON saved: {JSON_OUT}")

    # Generate PDFs
    PDF_OUT_DIR.mkdir(parents=True, exist_ok=True)
    EQUIP_PDF_DIR.mkdir(parents=True, exist_ok=True)

    pdf_count = 0
    for rec in records:
        inst = rec["instruction"]
        wo = inst.get("wo_number", "")
        location = inst.get("location", "")
        line, product = extract_line_product(location)
        # WO-20250113-001 → WO-20250113-001_A_AX-01.pdf
        pdf_name = f"{wo}_{line}_{product}.pdf"
        out_path = PDF_OUT_DIR / pdf_name
        generate_wo_pdf(rec, out_path)
        pdf_count += 1

    print(f"Generated {pdf_count} work order PDFs in {PDF_OUT_DIR}")

    # Tool list PDF
    tool_pdf_path = EQUIP_PDF_DIR / "TOOL_LIST_V1.pdf"
    generate_tool_list_pdf(tool_items, tool_pdf_path)
    print(f"Generated tool list PDF: {tool_pdf_path}")

    print(f"\n=== Total PDFs: {pdf_count + 1} ===")


if __name__ == "__main__":
    main()
